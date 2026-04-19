import re
import shutil
import json
import logging
from pathlib import Path
from typing import Any, cast

from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.worksheet.worksheet import Worksheet

from .name_clustering import canonicalize_name_list, cluster_names, person_sort_key
from .name_clustering_models import NameStringRecord
from .pricing import PRICING_OBJECT_NAME
from .utils import clone_or_copy_file, coerce_json_list, parse_sort_year, slugify, unique_copy_name

LOGGER = logging.getLogger(__name__)

_MONTH_NUMBERS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}
_MONTH_PATTERN = "|".join(sorted(_MONTH_NUMBERS, key=len, reverse=True))
_EXACT_DATE_PATTERN = re.compile(
    rf"\b(?P<month>{_MONTH_PATTERN})\.?\s+"
    r"(?P<day>\d{1,2})(?:st|nd|rd|th)?(?:,)?\s+"
    r"(?P<year>(?:18|19|20)\d{2})\b",
    re.IGNORECASE,
)
_MONTH_YEAR_PATTERN = re.compile(
    rf"\b(?P<month>{_MONTH_PATTERN})\.?\s+(?P<year>(?:18|19|20)\d{{2}})\b",
    re.IGNORECASE,
)
_QUALIFIED_YEAR_PATTERN = re.compile(
    r"\b(?P<qualifier>early|mid|late)[-\s]+(?P<year>(?:18|19|20)\d{2})\b",
    re.IGNORECASE,
)
_YEAR_PATTERN = re.compile(r"\b(?P<year>(?:18|19|20)\d{2})\b")
_UNSORTED_DATE = (9999, 12, 31)
_QUALIFIER_MONTHS = {
    "early": 2,
    "mid": 6,
    "late": 10,
}


def _safe_excel(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def _safe_display_list(values: list[str]) -> str:
    return "; ".join(value for value in values if value)


def _build_template_environment() -> Environment:
    return Environment(
        loader=FileSystemLoader(Path(__file__).parent / "templates"),
        autoescape=select_autoescape(["html", "xml"]),
        trim_blocks=True,
        lstrip_blocks=True,
    )


def _overlaps_existing(span: tuple[int, int], occupied: list[tuple[int, int]]) -> bool:
    return any(span[0] < other_end and other_start < span[1] for other_start, other_end in occupied)


def _extract_date_candidates(text: str) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    occupied: list[tuple[int, int]] = []

    for match in _EXACT_DATE_PATTERN.finditer(text):
        month = _MONTH_NUMBERS[match.group("month").casefold().rstrip(".")]
        day = int(match.group("day"))
        year = int(match.group("year"))
        occupied.append(match.span())
        candidates.append(
            {
                "sort_key": (year, month, day),
                "label": f"{match.group('month').rstrip('.').title()} {day}, {year}",
            }
        )

    for match in _MONTH_YEAR_PATTERN.finditer(text):
        if _overlaps_existing(match.span(), occupied):
            continue
        month = _MONTH_NUMBERS[match.group("month").casefold().rstrip(".")]
        year = int(match.group("year"))
        occupied.append(match.span())
        candidates.append(
            {
                "sort_key": (year, month, 1),
                "label": f"{match.group('month').rstrip('.').title()} {year}",
            }
        )

    for match in _QUALIFIED_YEAR_PATTERN.finditer(text):
        if _overlaps_existing(match.span(), occupied):
            continue
        qualifier = match.group("qualifier").casefold()
        year = int(match.group("year"))
        occupied.append(match.span())
        candidates.append(
            {
                "sort_key": (year, _QUALIFIER_MONTHS[qualifier], 1),
                "label": f"{qualifier} {year}",
            }
        )

    for match in _YEAR_PATTERN.finditer(text):
        if _overlaps_existing(match.span(), occupied):
            continue
        year = int(match.group("year"))
        candidates.append({"sort_key": (year, 1, 1), "label": str(year)})

    return candidates


def _best_date_metadata(
    values: list[str],
    fallback_text: str = "",
) -> tuple[tuple[int, int, int], str, str]:
    text = " ".join(value for value in values if value)
    if fallback_text:
        text = f"{text} {fallback_text}".strip()
    candidates = _extract_date_candidates(text)
    if not candidates:
        return _UNSORTED_DATE, "Undated", "Undated"

    candidates.sort(key=lambda candidate: (candidate["sort_key"], candidate["label"].casefold()))
    distinct_candidates: list[dict[str, Any]] = []
    seen_labels: set[str] = set()
    for candidate in candidates:
        label_key = candidate["label"].casefold()
        if label_key in seen_labels:
            continue
        seen_labels.add(label_key)
        distinct_candidates.append(candidate)

    best = distinct_candidates[0]
    if len(distinct_candidates) == 1:
        return best["sort_key"], best["label"], best["label"]
    return (
        best["sort_key"],
        best["label"],
        f"{distinct_candidates[0]['label']} -> {distinct_candidates[1]['label']}",
    )


def _truncate_text(text: str, limit: int = 105) -> str:
    value = " ".join((text or "").split())
    if len(value) <= limit:
        return value
    return value[: limit - 3].rstrip() + "..."


def _key_person_for_row(row: dict[str, Any]) -> str:
    if row["people"]:
        return row["people"][0]
    return "No named person"


def _page_label(page_start: int | None, page_end: int | None) -> str:
    if not isinstance(page_start, int) or page_start < 1:
        return "Pages unknown"
    if isinstance(page_end, int) and page_end >= page_start and page_end != page_start:
        return f"Pages {page_start}-{page_end}"
    return f"Page {page_start}"


def _pdf_page_href(report_href: str | None, page_start: int | None) -> str | None:
    if not report_href:
        return None
    if not isinstance(page_start, int) or page_start < 1:
        return report_href
    return f"{report_href}#page={page_start}"


def _format_time_period(mentions: list[dict[str, Any]]) -> str:
    if not mentions:
        return "Undated"
    first_label = mentions[0]["date_label"]
    last_label = mentions[-1]["date_label"]
    if first_label == last_label:
        return first_label
    return f"{first_label} to {last_label}"


def _pluralize(value: int, singular: str) -> str:
    suffix = "" if value == 1 else "s"
    return f"{value} {singular}{suffix}"


def _errors_label(*, failure_count: int, unanalyzed_count: int) -> str:
    parts: list[str] = []
    if failure_count > 0:
        parts.append(_pluralize(failure_count, "failure"))
    if unanalyzed_count > 0:
        parts.append(_pluralize(unanalyzed_count, "unanalyzed"))
    if not parts:
        return "none"
    return ", ".join(parts)


def _evidence_sort_key(row: dict[str, Any]) -> tuple[tuple[int, int, int], int, str, str]:
    page = row["page_start"] if isinstance(row["page_start"], int) else 10**9
    return (row["sort_key"], page, row["source_filename"].casefold(), row["summary"].casefold())


def _build_name_string_records(
    evidence_rows: list[dict[str, Any]],
    document_rows: list[dict[str, Any]],
    synthesis_payload: dict[str, Any] | None,
) -> list[NameStringRecord]:
    grouped: dict[str, dict[str, Any]] = {}

    def ensure_bucket(name: str) -> dict[str, Any]:
        return grouped.setdefault(
            name,
            {
                "mentions": 0,
                "source_filenames": set(),
                "context_notes": [],
                "contexts_seen": set(),
                "mentions_for_period": [],
            },
        )

    for row in evidence_rows:
        seen_in_row: set[str] = set()
        for raw_name in row["raw_people"]:
            cleaned = raw_name.strip()
            if not cleaned or cleaned in seen_in_row:
                continue
            seen_in_row.add(cleaned)
            bucket = ensure_bucket(cleaned)
            bucket["mentions"] += 1
            if row.get("source_filename"):
                bucket["source_filenames"].add(row["source_filename"])
            context_note = row.get("brief_summary") or row.get("summary") or ""
            if context_note and context_note not in bucket["contexts_seen"]:
                bucket["contexts_seen"].add(context_note)
                if len(bucket["context_notes"]) < 5:
                    bucket["context_notes"].append(context_note)
            bucket["mentions_for_period"].append(
                {
                    "sort_key": row["sort_key"],
                    "date_label": row["date_label"],
                }
            )

    for row in document_rows:
        if row.get("responsive") != 1:
            continue
        for raw_name in coerce_json_list(row.get("people_json")):
            cleaned = raw_name.strip()
            if not cleaned:
                continue
            bucket = ensure_bucket(cleaned)
            if row.get("canonical_filename"):
                bucket["source_filenames"].add(row["canonical_filename"])
            context_note = row.get("summary") or ""
            if context_note and context_note not in bucket["contexts_seen"]:
                bucket["contexts_seen"].add(context_note)
                if len(bucket["context_notes"]) < 5:
                    bucket["context_notes"].append(context_note)

    if synthesis_payload is not None:
        for raw_name in synthesis_payload.get("people", []):
            cleaned = raw_name.strip()
            if not cleaned:
                continue
            bucket = ensure_bucket(cleaned)
            context_note = synthesis_payload.get("answer") or ""
            if context_note and context_note not in bucket["contexts_seen"]:
                bucket["contexts_seen"].add(context_note)
                if len(bucket["context_notes"]) < 5:
                    bucket["context_notes"].append(context_note)

    records: list[NameStringRecord] = []
    for index, name in enumerate(sorted(grouped, key=str.casefold), start=1):
        bucket = grouped[name]
        mentions_for_period = sorted(
            bucket["mentions_for_period"],
            key=lambda mention: (mention["sort_key"], mention["date_label"].casefold()),
        )
        time_period = _format_time_period(mentions_for_period) if mentions_for_period else None
        records.append(
            NameStringRecord(
                id=index,
                name_string=name,
                mentions=int(bucket["mentions"]),
                time_period=time_period,
                source_filenames=tuple(sorted(bucket["source_filenames"], key=str.casefold)),
                context_notes=tuple(bucket["context_notes"]),
            )
        )
    return records


def _build_authoritative_name_mapping(
    *,
    evidence_rows: list[dict[str, Any]],
    document_rows: list[dict[str, Any]],
    synthesis_payload: dict[str, Any] | None,
    name_clustering_method: str,
    model_name: str,
    allow_gemini: bool,
) -> dict[str, str]:
    records = _build_name_string_records(evidence_rows, document_rows, synthesis_payload)
    if not records:
        return {}
    clustering_result = cluster_names(
        records,
        method=name_clustering_method,
        model_name=model_name,
        allow_gemini=allow_gemini,
    )
    record_name_by_id = {record.id: record.name_string for record in records}
    canonical_by_raw: dict[str, str] = {}
    for cluster in clustering_result.clusters:
        for member_id in cluster.member_name_ids:
            canonical_by_raw[record_name_by_id[member_id]] = cluster.representative_name
    LOGGER.info(
        "Applied %s name clustering across %s unique extracted name strings, producing %s authoritative names.",
        clustering_result.method,
        len(records),
        len(clustering_result.clusters),
    )
    return canonical_by_raw


def _build_people_index(evidence_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in evidence_rows:
        matched_names_by_person: dict[str, list[str]] = {}
        for raw_person in row["raw_people"]:
            cleaned = raw_person.strip()
            if not cleaned:
                continue
            canonical_person = row["canonical_name_by_raw"].get(cleaned, cleaned)
            matched_names = matched_names_by_person.setdefault(canonical_person, [])
            if cleaned not in matched_names:
                matched_names.append(cleaned)
        for canonical_person, matched_names in matched_names_by_person.items():
            grouped.setdefault(canonical_person, []).append(
                {
                    "anchor": row["anchor"],
                    "date_label": row["date_label"],
                    "date_badge_label": row["date_badge_label"],
                    "sort_key": row["sort_key"],
                    "brief_summary": row["brief_summary"],
                    "source_filename": row["source_filename"],
                    "page_start": row["page_start"],
                    "page_label": row["page_label"],
                    "page_href": _pdf_page_href(row.get("report_href"), row["page_start"]),
                    "people": row["people"],
                    "matched_names": matched_names,
                }
            )

    used_toggle_ids: set[str] = set()
    people_index: list[dict[str, Any]] = []
    for person, mentions in grouped.items():
        mentions.sort(
            key=lambda mention: (
                mention["sort_key"],
                mention["page_start"] if isinstance(mention.get("page_start"), int) else 10**9,
                mention["source_filename"].casefold(),
                mention["brief_summary"].casefold(),
            )
        )
        toggle_id = unique_copy_name(f"person-{slugify(person)}", used_toggle_ids)
        people_index.append(
            {
                "person": person,
                "mention_count": len(mentions),
                "time_period": _format_time_period(mentions),
                "mentions": mentions,
                "toggle_id": toggle_id,
            }
        )

    people_index.sort(key=lambda row: person_sort_key(row["person"]))
    return people_index


def _build_responsive_documents_index(
    document_rows: list[dict[str, Any]],
    evidence_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    rows_by_sha: dict[str, list[dict[str, Any]]] = {}
    for row in evidence_rows:
        rows_by_sha.setdefault(row["document_sha256"], []).append(row)

    responsive_documents: list[dict[str, Any]] = []
    for document_row in document_rows:
        if document_row.get("responsive") != 1:
            continue
        evidence_items = rows_by_sha.get(document_row["sha256"], [])
        if not evidence_items:
            continue
        evidence_items.sort(key=_evidence_sort_key)
        responsive_documents.append(
            {
                "document_sha256": document_row["sha256"],
                "anchor": f"document-{slugify(document_row['canonical_filename'])}-{document_row['sha256'][:8]}",
                "source_filename": document_row["canonical_filename"],
                "page_count": int(document_row.get("stored_page_count") or 0),
                "document_summary": document_row.get("summary") or "",
                "report_href": evidence_items[0].get("report_href"),
                "evidence_items": [
                    {
                        "anchor": item["anchor"],
                        "page_label": item["page_label"],
                        "page_href": item.get("page_href"),
                        "people": item["people"],
                        "brief_summary": item["brief_summary"],
                    }
                    for item in evidence_items
                ],
            }
        )
    responsive_documents.sort(key=lambda row: row["source_filename"].casefold())
    return responsive_documents


def generate_reports(
    *,
    db,
    query_id: int,
    output_dir: Path,
    question: str,
    project_name: str,
    model_name: str,
    name_clustering_method: str,
    allow_gemini: bool,
    run_summary: dict[str, Any] | None = None,
) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    pdf_dir = output_dir / "pdfs"
    if pdf_dir.exists():
        shutil.rmtree(pdf_dir)
    pdf_dir.mkdir(parents=True, exist_ok=True)

    document_rows = [dict(row) for row in db.report_document_rows(query_id)]
    evidence_rows_raw = [dict(row) for row in db.report_evidence_rows(query_id)]
    failure_rows = [dict(row) for row in db.report_failure_rows(query_id)]
    synthesis_row = db.fetch_synthesis(query_id)
    pricing_snapshot_row = db.fetch_latest_object(PRICING_OBJECT_NAME)
    if pricing_snapshot_row is not None:
        db.backfill_missing_usage_costs(
            query_id=query_id,
            pricing_snapshot=json.loads(pricing_snapshot_row["json_object"]),
        )
    usage_summary = db.calculate_query_usage_summary(query_id)

    responsive_sha256s = {row["document_sha256"] for row in evidence_rows_raw}
    responsive_sha256s.update(row["sha256"] for row in document_rows if row.get("responsive") == 1)

    used_names: set[str] = set()
    linked_files: dict[str, str] = {}
    for row in document_rows:
        if row["sha256"] not in responsive_sha256s:
            continue
        source_path = row.get("source_path")
        if not source_path:
            continue
        source = Path(source_path)
        if not source.exists():
            continue
        report_name = unique_copy_name(row["canonical_filename"], used_names)
        target = pdf_dir / report_name
        clone_or_copy_file(source, target)
        linked_files[row["sha256"]] = f"pdfs/{report_name}"

    evidence_rows: list[dict[str, Any]] = []
    for row in evidence_rows_raw:
        dates = coerce_json_list(row["dates_json"])
        raw_people = coerce_json_list(row["people_json"])
        places = coerce_json_list(row["places_json"])
        sort_key, date_label, date_badge_label = _best_date_metadata(
            dates,
            row["evidence_summary"] or row["document_summary"] or "",
        )
        evidence_rows.append(
            {
                "document_sha256": row["document_sha256"],
                "source_filename": row["source_filename"],
                "source_path": row["source_path"],
                "document_summary": row["document_summary"] or "",
                "relevance_score": row["relevance_score"],
                "page_start": row["page_start"],
                "page_end": row["page_end"],
                "page_label": _page_label(row["page_start"], row["page_end"]),
                "summary": row["evidence_summary"],
                "brief_summary": _truncate_text(row["evidence_summary"] or row["document_summary"] or ""),
                "raw_people": raw_people,
                "people": raw_people,
                "places": places,
                "dates": dates,
                "date_label": date_label,
                "date_badge_label": date_badge_label,
                "sort_key": sort_key,
                "key_person": "",
                "anchor": "",
                "report_href": linked_files.get(row["document_sha256"]),
                "page_href": _pdf_page_href(
                    linked_files.get(row["document_sha256"]),
                    row["page_start"],
                ),
                "document_anchor": "",
            }
        )
    evidence_rows.sort(key=_evidence_sort_key)

    synthesis_payload = None
    if synthesis_row is not None and synthesis_row["status"] == "succeeded":
        synthesis_payload = {
            "answer": synthesis_row["answer"],
            "key_findings": coerce_json_list(synthesis_row["key_findings_json"]),
            "people": coerce_json_list(synthesis_row["people_json"]),
            "places": coerce_json_list(synthesis_row["places_json"]),
            "dates": coerce_json_list(synthesis_row["dates_json"]),
            "reasoning_notes": synthesis_row["reasoning_notes"],
        }

    authoritative_name_by_raw = _build_authoritative_name_mapping(
        evidence_rows=evidence_rows,
        document_rows=document_rows,
        synthesis_payload=synthesis_payload,
        name_clustering_method=name_clustering_method,
        model_name=model_name,
        allow_gemini=allow_gemini,
    )
    for row in evidence_rows:
        row["canonical_name_by_raw"] = authoritative_name_by_raw
        row["people"] = canonicalize_name_list(row["raw_people"], authoritative_name_by_raw)
    for index, row in enumerate(evidence_rows, start=1):
        row["anchor"] = f"evidence-{index}"
        row["key_person"] = _key_person_for_row(row)
    responsive_documents_index = _build_responsive_documents_index(document_rows, evidence_rows)
    document_anchor_map = {
        row["document_sha256"]: row["anchor"] for row in responsive_documents_index
    }
    for row in evidence_rows:
        row["document_anchor"] = document_anchor_map.get(row["document_sha256"], "")

    for row in document_rows:
        row["people"] = canonicalize_name_list(
            coerce_json_list(row.get("people_json")),
            authoritative_name_by_raw,
        )
        row["places"] = coerce_json_list(row.get("places_json"))
        row["dates"] = coerce_json_list(row.get("dates_json"))

    if synthesis_payload is not None:
        synthesis_payload["people"] = canonicalize_name_list(
            synthesis_payload["people"],
            authoritative_name_by_raw,
        )

    people_index = _build_people_index(evidence_rows)
    unanalyzed_rows = [row for row in document_rows if row.get("status") != "succeeded"]

    total_pages_scanned = (
        int(run_summary["scanned_pages"])
        if run_summary and run_summary.get("scanned_pages") is not None
        else sum(int(row.get("stored_page_count") or 0) for row in document_rows)
    )
    summary = {
        "total_documents": len(document_rows),
        "total_pages_scanned": total_pages_scanned,
        "responsive_evidence_rows": len(evidence_rows),
        "responsive_documents": sum(1 for row in document_rows if row.get("responsive") == 1),
        "failed_documents": len(failure_rows),
        "unanalyzed_documents": len(unanalyzed_rows),
        "errors_label": _errors_label(
            failure_count=len(failure_rows),
            unanalyzed_count=len(unanalyzed_rows),
        ),
        "responsive_document_index_count": len(responsive_documents_index),
    }
    if run_summary is not None:
        run_summary = {
            **run_summary,
            "prompt_tokens": usage_summary["prompt_tokens"],
            "candidate_tokens": usage_summary["candidate_tokens"],
            "total_tokens": usage_summary["total_tokens"],
            "input_cost_usd": usage_summary["input_cost_usd"],
            "output_cost_usd": usage_summary["output_cost_usd"],
            "total_cost_usd": usage_summary["total_cost_usd"],
            "pricing_available": pricing_snapshot_row is not None,
        }

    env = _build_template_environment()
    template = env.get_template("report.html.jinja")
    html = template.render(
        title=project_name,
        question=question,
        summary=summary,
        run_summary=run_summary,
        synthesis=synthesis_payload,
        evidence_rows=evidence_rows,
        responsive_documents_index=responsive_documents_index,
        people_index=people_index,
        document_rows=document_rows,
        unanalyzed_rows=unanalyzed_rows,
        failure_rows=failure_rows,
    )
    html_path = output_dir / "report.html"
    html_path.write_text(html.rstrip() + "\n", encoding="utf-8")

    workbook = Workbook()
    evidence_sheet = workbook.active
    assert evidence_sheet is not None
    evidence_sheet = cast(Worksheet, evidence_sheet)
    evidence_sheet.title = "Evidence"
    evidence_sheet.append(
        [
            "sort_date",
            "sort_year",
            "source_filename",
            "page_start",
            "page_end",
            "page_label",
            "key_person",
            "summary",
            "people",
            "places",
            "dates",
            "document_summary",
            "relevance_score",
            "source_path",
        ]
    )
    for row in evidence_rows:
        evidence_sheet.append(
            [
                _safe_excel(row["date_label"]),
                _safe_excel(parse_sort_year(row["dates"], row["summary"])),
                _safe_excel(row["source_filename"]),
                _safe_excel(row["page_start"]),
                _safe_excel(row["page_end"]),
                _safe_excel(row["page_label"]),
                _safe_excel(row["key_person"]),
                _safe_excel(row["summary"]),
                _safe_excel(_safe_display_list(row["people"])),
                _safe_excel(_safe_display_list(row["places"])),
                _safe_excel(_safe_display_list(row["dates"])),
                _safe_excel(row["document_summary"]),
                _safe_excel(row["relevance_score"]),
                _safe_excel(row["source_path"]),
            ]
        )

    documents_sheet = workbook.create_sheet("Documents")
    documents_sheet.append(
        [
            "source_filename",
            "relative_path",
            "status",
            "responsive",
            "relevance_score",
            "summary",
            "people",
            "places",
            "dates",
            "evidence_count",
            "failure_type",
            "error_text",
        ]
    )
    for row in document_rows:
        documents_sheet.append(
            [
                _safe_excel(row["canonical_filename"]),
                _safe_excel(row.get("relative_path")),
                _safe_excel(row.get("status")),
                _safe_excel(row.get("responsive")),
                _safe_excel(row.get("relevance_score")),
                _safe_excel(row.get("summary")),
                _safe_excel(_safe_display_list(row.get("people", []))),
                _safe_excel(_safe_display_list(row.get("places", []))),
                _safe_excel(_safe_display_list(row.get("dates", []))),
                _safe_excel(row.get("evidence_count")),
                _safe_excel(row.get("failure_type")),
                _safe_excel(row.get("error_text")),
            ]
        )

    failures_sheet = workbook.create_sheet("Failures")
    failures_sheet.append(
        ["source_filename", "status", "failure_type", "error_text", "source_path", "completed_at"]
    )
    for row in failure_rows:
        failures_sheet.append(
            [
                _safe_excel(row.get("source_filename")),
                _safe_excel(row.get("status")),
                _safe_excel(row.get("failure_type")),
                _safe_excel(row.get("error_text")),
                _safe_excel(row.get("source_path")),
                _safe_excel(row.get("completed_at")),
            ]
        )

    xlsx_path = output_dir / "report.xlsx"
    workbook.save(xlsx_path)
    return html_path, xlsx_path
